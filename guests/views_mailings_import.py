import re
from io import BytesIO
from itertools import chain

from django.contrib import messages
from django.db import transaction
from django.views import View
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook

from .forms import MailingImportPhonesForm
from .models import Guest, InteractionButtonSet, Mailing, MailingGuest
from guests.services.mailing_import_audience import (
    build_mailing_import_audience_selection,
)
from guests.services.mailing_reverse_import import (
    MAILING_IMPORT_OPERATION_EXCLUDE,
    MAILING_IMPORT_OPERATION_INCLUDE,
    ReverseMailingImportError,
    build_reverse_mailing_import_calculation,
    confirm_reverse_mailing_import,
    create_reverse_import_preview_token,
    read_reverse_exclusion_xlsx,
)
from guests.services.template_render import render_message_for_guest


def normalize_phone(raw: str) -> str | None:
    """
    Приводим телефон к единому виду для поиска.
    Простейшая нормализация: берём цифры, оставляем последние 10.
    """
    if not raw:
        return None
    digits = re.sub(r"\D+", "", str(raw))
    if not digits:
        return None

    # часто в РФ номера 11 цифр (7/8 + 10 цифр)
    if len(digits) >= 10:
        digits10 = digits[-10:]
        return digits10  # храним/сравниваем по последним 10 цифрам

    return None


def normalize_header(raw: str) -> str:
    return str(raw or "").strip().lower().replace(" ", "_")


def read_recipients_from_xlsx(file_obj) -> list[dict[str, str]]:
    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows) or ()
        except StopIteration:
            return []

        headers = {
            normalize_header(value): index
            for index, value in enumerate(first_row)
            if normalize_header(value)
        }
        phone_index = None
        telegram_index = None
        for header_name in ("phone", "телефон", "phone_e164", "phone_number"):
            if header_name in headers:
                phone_index = headers[header_name]
                break
        for header_name in ("telegram_external_id", "external_id", "telegram_id", "telegram_chat_id"):
            if header_name in headers:
                telegram_index = headers[header_name]
                break

        has_header = phone_index is not None
        if phone_index is None:
            phone_index = 0
        if telegram_index is None and len(first_row) > 1:
            telegram_index = 1

        recipients: list[dict[str, str]] = []
        data_rows = rows if has_header else chain((first_row,), rows)
        for row in data_rows:
            if not row:
                continue
            phone_raw = row[phone_index] if phone_index < len(row) else None
            phone = normalize_phone(phone_raw)
            if not phone:
                continue
            telegram_external_id = ""
            if telegram_index is not None and telegram_index < len(row):
                telegram_external_id = str(row[telegram_index] or "").strip()
            recipients.append(
                {
                    "phone": phone,
                    "telegram_external_id": telegram_external_id[:32],
                }
            )

        return recipients
    finally:
        wb.close()


def resolve_next_url(request, fallback_url: str) -> str:
    """
    Безопасно вернуть URL для redirect после импорта.

    Позволяет new UI передавать `next`, не ломая legacy поведение.
    """
    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if next_url and url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback_url


class MailingImportPhonesView(View):
    """
    POST: загрузить Excel, сопоставить телефоны с Guest, добавить в MailingGuest.
    """
    def post(self, request, pk: int):
        mailing = get_object_or_404(Mailing, pk=pk)
        fallback_url = reverse("mailing_edit", args=[mailing.id])
        redirect_url = resolve_next_url(request, fallback_url)
        form = MailingImportPhonesForm(request.POST, request.FILES)

        if not form.is_valid():
            # возвращаем на форму редактирования (или куда тебе удобнее)
            request.session["mailing_import_error"] = str(form.errors)
            return redirect(redirect_url)

        operation = form.cleaned_data["import_operation"]
        if operation == MAILING_IMPORT_OPERATION_EXCLUDE:
            return self._handle_reverse_import(
                request=request,
                mailing=mailing,
                form=form,
                redirect_url=redirect_url,
            )
        return self._handle_include_import(
            request=request,
            mailing=mailing,
            form=form,
            redirect_url=redirect_url,
        )

    def _handle_reverse_import(self, *, request, mailing, form, redirect_url):
        """Выполняет безопасный предпросмотр либо подтверждение комплемента."""

        try:
            excel = read_reverse_exclusion_xlsx(form.cleaned_data["file"])
        except ReverseMailingImportError as exc:
            return self._redirect_with_error(request, redirect_url, str(exc))

        audience_group = form.cleaned_data["audience_channel_group"]
        import_action = str(request.POST.get("import_action") or "preview").strip().lower()
        if import_action == "confirm":
            try:
                _locked_mailing, calculation, created_count = confirm_reverse_mailing_import(
                    mailing_id=int(mailing.id),
                    excel=excel,
                    audience_group=audience_group,
                    preview_token=str(request.POST.get("preview_token") or ""),
                )
            except ReverseMailingImportError as exc:
                return self._redirect_with_error(request, redirect_url, str(exc))

            request.session["mailing_import_report"] = calculation.public_report(
                added=created_count
            )
            messages.success(
                request,
                f"Обратная аудитория подтверждена: добавлено гостей {created_count}.",
            )
            return redirect(redirect_url)

        calculation = build_reverse_mailing_import_calculation(
            mailing=mailing,
            excel=excel,
            audience_group=audience_group,
        )
        preview_token = (
            create_reverse_import_preview_token(calculation)
            if calculation.can_confirm
            else ""
        )
        return render(
            request,
            "mailing/import_reverse_confirm.html",
            {
                "mailing": mailing,
                "preview": calculation.public_report(),
                "preview_token": preview_token,
                "next_url": redirect_url,
            },
        )

    def _handle_include_import(self, *, request, mailing, form, redirect_url):
        """Сохраняет прежний прямой импорт значением по умолчанию."""

        try:
            recipients_raw = read_recipients_from_xlsx(form.cleaned_data["file"])
        except Exception:
            return self._redirect_with_error(
                request,
                redirect_url,
                "Не удалось прочитать Excel. Проверьте, что это исправный файл .xlsx.",
            )

        total_loaded = len(recipients_raw)
        recipients_by_phone: dict[str, dict[str, str]] = {}
        duplicate_rows = 0
        for recipient in recipients_raw:
            phone = recipient["phone"]
            if phone in recipients_by_phone:
                duplicate_rows += 1
                if not recipients_by_phone[phone].get("telegram_external_id") and recipient.get("telegram_external_id"):
                    recipients_by_phone[phone] = recipient
                continue
            recipients_by_phone[phone] = recipient
        phones_unique = sorted(recipients_by_phone)

        # 1) находим гостей по телефону
        # ВАЖНО: в БД у тебя phone может быть с +7, пробелами и т.п.
        # Поэтому делаем "поиск по цифрам": берём гостей, у кого phone содержит последние 10.
        # Это не супер быстро на миллионах, но для начала ок.
        guests_found = []
        not_found = []

        # Быстрее: одним запросом не получится идеально из-за normalize,
        # поэтому делаем так: берём всех гостей по совпадению "окончания".
        # (Если у тебя Postgres — можно сделать regexp_replace на уровне SQL, но это усложнение.)
        candidates = Guest.objects.filter(phone__isnull=False).exclude(phone__exact="")

        phone_to_guest = {}
        for g in candidates.iterator(chunk_size=5000):
            gp = normalize_phone(g.phone)
            if gp:
                phone_to_guest.setdefault(gp, g)

        for p in phones_unique:
            g = phone_to_guest.get(p)
            if g:
                guests_found.append(g)
            else:
                not_found.append(p)

        found_count = len(guests_found)

        # 2) Выясняем, кто уже есть в рассылке: пара mailing+guest уникальна.
        found_ids = [g.id for g in guests_found]
        already_ids = set(
            MailingGuest.objects.filter(mailing=mailing, guest_id__in=found_ids)
            .values_list("guest_id", flat=True)
        )

        guest_external_ids = {
            g.id: recipients_by_phone.get(normalize_phone(g.phone) or "", {}).get(
                "telegram_external_id",
                "",
            )
            for g in guests_found
        }

        # 2.5) Используем тот же планировщик, который применяется при фактической
        # постановке рассылки в очередь. Это исключает расхождение между импортом
        # и отправкой для новых и исторических каналов.
        selected_bot_ids = list(
            mailing.bot_profiles.filter(is_active=True).values_list("id", flat=True)
        )
        audience_selection = build_mailing_import_audience_selection(
            found_ids,
            selected_bot_ids=selected_bot_ids,
            target_mode=mailing.target_mode,
            audience_group=form.cleaned_data["audience_channel_group"],
            telegram_external_ids=guest_external_ids,
        )
        selected_historical_guest_ids = (
            audience_selection.selected_guest_ids
            & audience_selection.historical_guest_ids
        )
        if selected_historical_guest_ids and mailing.button_set != InteractionButtonSet.NONE:
            messages.error(
                request,
                (
                    "Выбранная аудитория содержит исторические Telegram-маршруты, "
                    "которые нельзя добавить в рассылку с кнопками. Сначала "
                    "выберите вариант «Без кнопок»."
                ),
            )
            return redirect(redirect_url)
        to_add = [
            guest
            for guest in guests_found
            if guest.id in audience_selection.selected_guest_ids
            and guest.id not in already_ids
        ]

        # 3) создаём строки MailingGuest
        now = timezone.now()
        template_text = mailing.template.message_text
        scheduled_dt = mailing.scheduled_time_begin  # логика: отправлять можно не раньше начала окна

        rows = []
        for g in to_add:
            rendered_text = render_message_for_guest(template_text, g)
            rows.append(MailingGuest(
                mailing=mailing,
                guest=g,
                phone=g.phone,
                email=g.email,
                text_mailing_list=rendered_text,
                scheduled_datetime=scheduled_dt,
                status=MailingGuest.Status.PLANNED,
                external_id=(
                    guest_external_ids.get(g.id) or None
                    if g.id in audience_selection.file_telegram_external_id_guest_ids
                    else None
                ),
                created_at=now,
            ))

        with transaction.atomic():
            MailingGuest.objects.bulk_create(rows)
            source_snapshot = dict(mailing.source_filter_snapshot or {})
            import_groups = source_snapshot.get("mailing_import_audience_groups") or []
            if not isinstance(import_groups, (list, tuple, set)):
                import_groups = []
            normalized_groups = {
                str(value or "").strip() for value in import_groups if str(value or "").strip()
            }
            normalized_groups.add(audience_selection.audience_group)
            source_snapshot["mailing_import_audience_groups"] = sorted(normalized_groups)
            source_snapshot["mailing_import_contains_historical"] = bool(
                source_snapshot.get("mailing_import_contains_historical")
                or selected_historical_guest_ids
            )
            mailing.source_filter_snapshot = source_snapshot
            mailing.save(update_fields=["source_filter_snapshot"])

        added_count = len(rows)
        already_count = len(
            already_ids.intersection(audience_selection.selected_guest_ids)
        )
        not_found_count = len(not_found)
        legacy_external_id_count = sum(1 for row in rows if row.external_id)

        # 4) сохраняем отчёт (проще всего в session, чтобы показать на странице)
        request.session["mailing_import_report"] = {
            "import_operation": MAILING_IMPORT_OPERATION_INCLUDE,
            "import_operation_label": "Добавить гостей, указанных в Excel",
            "total_loaded": total_loaded,
            "unique_phones": len(phones_unique),
            "duplicate_rows": duplicate_rows,
            "found": found_count,
            "audience_channel_group": audience_selection.audience_group,
            "audience_channel_group_label": audience_selection.audience_group_label,
            "sendable_new_bots": len(audience_selection.new_bot_guest_ids),
            "sendable_historical": len(audience_selection.historical_guest_ids),
            "sendable_total": len(audience_selection.sendable_guest_ids),
            "excluded_without_channel": len(
                audience_selection.without_sendable_channel_guest_ids
            ),
            "excluded_by_audience_group": len(
                audience_selection.excluded_by_audience_group_guest_ids
            ),
            "added": added_count,
            "already": already_count,
            "not_found": not_found_count,
            "legacy_external_id": legacy_external_id_count,
        }

        return redirect(redirect_url)

    @staticmethod
    def _redirect_with_error(request, redirect_url: str, error: str):
        messages.error(request, error)
        request.session["mailing_import_error"] = error
        return redirect(redirect_url)


class MailingImportTemplateDownloadView(View):
    """
    Скачать шаблон Excel с одним столбцом phone.
    """
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "phones"
        ws["A1"] = "phone"
        ws["B1"] = "telegram_external_id"
        ws["A2"] = "+79991234567"
        ws["B2"] = "123456789"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="mailing_phones_template.xlsx"'
        return resp
