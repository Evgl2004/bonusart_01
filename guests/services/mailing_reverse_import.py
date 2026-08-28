"""Безопасный предварительный расчёт обратного Excel-импорта рассылки."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from typing import Any, Iterable

from django.core import signing
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from guests.models import (
    BotProfile,
    DispatchTask,
    Guest,
    GuestBotBinding,
    HistoricalTelegramChannel,
    InteractionButtonSet,
    Mailing,
    MailingGuest,
    VtelemaxRecipientChannel,
)
from guests.services.guest_resolution import normalize_phone10
from guests.services.mailing_delivery_targets import (
    CHANNEL_MODE_BINDING,
    CHANNEL_MODE_HISTORICAL_TELEGRAM,
    CHANNEL_MODE_LEGACY_TELEGRAM,
    MailingDeliveryRow,
    build_mailing_delivery_plan,
)
from guests.services.mailing_import_audience import (
    MAILING_IMPORT_AUDIENCE_ALL_SENDABLE,
    MAILING_IMPORT_AUDIENCE_HISTORICAL_TELEGRAM,
    MAILING_IMPORT_AUDIENCE_LABELS,
    MAILING_IMPORT_AUDIENCE_NEW_BOTS,
    normalize_mailing_import_audience,
)
from guests.services.template_render import render_message_for_guest


MAILING_IMPORT_OPERATION_INCLUDE = "include"
MAILING_IMPORT_OPERATION_EXCLUDE = "exclude"
MAILING_IMPORT_OPERATION_CHOICES = (
    (MAILING_IMPORT_OPERATION_INCLUDE, "Добавить гостей, указанных в Excel"),
    (
        MAILING_IMPORT_OPERATION_EXCLUDE,
        "Добавить всех допустимых гостей, кроме указанных в Excel",
    ),
)
MAILING_IMPORT_OPERATION_LABELS = dict(MAILING_IMPORT_OPERATION_CHOICES)

REVERSE_IMPORT_WARNING_PERCENT = 60
REVERSE_IMPORT_TOKEN_MAX_AGE_SECONDS = 30 * 60
REVERSE_IMPORT_SIGNING_SALT = "guests.mailing-reverse-import.v1"

_PHONE_HEADERS = ("phone", "телефон", "phone_e164", "phone_number")
_HISTORICAL_CHANNEL_MODES = {
    CHANNEL_MODE_HISTORICAL_TELEGRAM,
    CHANNEL_MODE_LEGACY_TELEGRAM,
}


class ReverseMailingImportError(ValueError):
    """Ожидаемая безопасная ошибка обратного импорта."""


class ReverseMailingImportPreviewMismatch(ReverseMailingImportError):
    """Предпросмотр больше не соответствует подтверждаемому расчёту."""


@dataclass(frozen=True)
class ReverseExcelAnalysis:
    """Результат потокового разбора Excel без записи исходных телефонов."""

    file_sha256: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_rows: int
    unique_phones: tuple[str, ...]


@dataclass(frozen=True)
class ReverseMailingImportCalculation:
    """Полный расчёт обратной аудитории и безопасная публичная сводка."""

    mailing_id: int
    audience_group: str
    audience_group_label: str
    file_sha256: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_rows: int
    unique_phones: int
    matched_phones: int
    unmatched_phones: int
    matched_guest_records: int
    source_deliverable_guests: int
    excluded_guest_records: int
    final_recipients: int
    final_share_percent: float
    guests_without_normalized_phone: int
    provider_telegram: int
    provider_vk: int
    provider_max: int
    historical_guests: int
    warning_large_audience: bool
    blockers: tuple[str, ...]
    warnings: tuple[str, ...]
    campaign_fingerprint: str
    audience_digest: str
    calculation_digest: str
    final_guest_ids: tuple[int, ...]

    @property
    def can_confirm(self) -> bool:
        return not self.blockers

    def public_report(self, *, added: int | None = None) -> dict[str, Any]:
        """Возвращает только агрегаты, допустимые для интерфейса и сессии."""

        report: dict[str, Any] = {
            "import_operation": MAILING_IMPORT_OPERATION_EXCLUDE,
            "import_operation_label": MAILING_IMPORT_OPERATION_LABELS[
                MAILING_IMPORT_OPERATION_EXCLUDE
            ],
            "audience_channel_group": self.audience_group,
            "audience_channel_group_label": self.audience_group_label,
            "file_sha256": self.file_sha256,
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "invalid_rows": self.invalid_rows,
            "duplicate_rows": self.duplicate_rows,
            "unique_phones": self.unique_phones,
            "matched_phones": self.matched_phones,
            "unmatched_phones": self.unmatched_phones,
            "matched_guest_records": self.matched_guest_records,
            "source_deliverable_guests": self.source_deliverable_guests,
            "excluded_guest_records": self.excluded_guest_records,
            "final_recipients": self.final_recipients,
            "final_share_percent": self.final_share_percent,
            "guests_without_normalized_phone": self.guests_without_normalized_phone,
            "provider_telegram": self.provider_telegram,
            "provider_vk": self.provider_vk,
            "provider_max": self.provider_max,
            "historical_guests": self.historical_guests,
            "warning_large_audience": self.warning_large_audience,
            "warning_percent": REVERSE_IMPORT_WARNING_PERCENT,
            "blockers": list(self.blockers),
            "warnings": list(self.warnings),
        }
        if added is not None:
            report["added"] = int(added)
        return report


def normalize_mailing_import_operation(value: str | None) -> str:
    """Сохраняет прямой импорт значением по умолчанию для старых запросов."""

    normalized = str(value or "").strip().lower()
    if normalized == MAILING_IMPORT_OPERATION_EXCLUDE:
        return MAILING_IMPORT_OPERATION_EXCLUDE
    return MAILING_IMPORT_OPERATION_INCLUDE


def read_reverse_exclusion_xlsx(file_obj) -> ReverseExcelAnalysis:
    """Потоково читает строгий Excel-список исключаемых телефонов."""

    file_sha256 = _hash_uploaded_file(file_obj)
    workbook = None
    try:
        file_obj.seek(0)
        workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
        except StopIteration as exc:
            raise ReverseMailingImportError("Excel не содержит строк.") from exc

        headers = {
            _normalize_header(value): index
            for index, value in enumerate(first_row or ())
            if _normalize_header(value)
        }
        phone_index = next(
            (headers[name] for name in _PHONE_HEADERS if name in headers),
            None,
        )
        if phone_index is None:
            raise ReverseMailingImportError(
                "В Excel не найден обязательный столбец phone (телефон)."
            )

        total_rows = 0
        valid_rows = 0
        invalid_rows = 0
        duplicate_rows = 0
        unique_phones: set[str] = set()

        for row in rows:
            total_rows += 1
            phone_raw = row[phone_index] if phone_index < len(row or ()) else None
            phone = normalize_phone10(phone_raw)
            if not phone:
                invalid_rows += 1
                continue
            valid_rows += 1
            if phone in unique_phones:
                duplicate_rows += 1
                continue
            unique_phones.add(phone)

        if total_rows == 0:
            raise ReverseMailingImportError("Excel не содержит строк с данными.")
        if not unique_phones:
            raise ReverseMailingImportError(
                "В Excel нет ни одного корректного телефона для исключения."
            )

        return ReverseExcelAnalysis(
            file_sha256=file_sha256,
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            duplicate_rows=duplicate_rows,
            unique_phones=tuple(sorted(unique_phones)),
        )
    except ReverseMailingImportError:
        raise
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise ReverseMailingImportError(
            "Не удалось прочитать Excel. Проверьте, что это исправный файл .xlsx."
        ) from exc
    except Exception as exc:
        # Повреждённый ZIP-контейнер openpyxl может поднимать исключения zipfile.
        raise ReverseMailingImportError(
            "Не удалось прочитать Excel. Проверьте, что это исправный файл .xlsx."
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
        try:
            file_obj.seek(0)
        except (AttributeError, OSError):
            pass


def build_reverse_mailing_import_calculation(
    *,
    mailing: Mailing,
    excel: ReverseExcelAnalysis,
    audience_group: str,
) -> ReverseMailingImportCalculation:
    """Строит доказуемый комплемент штатной технически доставляемой аудитории."""

    normalized_group = normalize_mailing_import_audience(audience_group)
    configured_bots = list(
        mailing.bot_profiles.order_by("id").values(
            "id", "code", "provider_type", "is_active"
        )
    )
    selected_bot_ids = [int(row["id"]) for row in configured_bots]
    candidate_guest_ids = _collect_delivery_candidate_guest_ids(configured_bots)

    delivery_plan = build_mailing_delivery_plan(
        candidate_guest_ids,
        selected_bot_ids=selected_bot_ids,
        target_mode=mailing.target_mode,
    )
    selected_rows = tuple(
        row for row in delivery_plan.rows if _row_belongs_to_group(row, normalized_group)
    )
    selected_row_by_guest_id = {int(row.guest_id): row for row in selected_rows}
    source_guest_ids = set(selected_row_by_guest_id)
    exclusion_phones = set(excel.unique_phones)

    normalized_phone_by_guest_id: dict[int, str] = {}
    matched_phones: set[str] = set()
    matched_guest_records = 0
    guest_rows = Guest.objects.order_by("id").values_list("id", "phone")
    for guest_id, raw_phone in guest_rows.iterator(chunk_size=5000):
        normalized_phone = normalize_phone10(raw_phone)
        if normalized_phone and int(guest_id) in source_guest_ids:
            normalized_phone_by_guest_id[int(guest_id)] = normalized_phone
        if normalized_phone in exclusion_phones:
            matched_phones.add(normalized_phone)
            matched_guest_records += 1

    excluded_guest_ids = {
        guest_id
        for guest_id in source_guest_ids
        if normalized_phone_by_guest_id.get(guest_id) in exclusion_phones
    }
    without_phone_guest_ids = {
        guest_id
        for guest_id in source_guest_ids
        if guest_id not in normalized_phone_by_guest_id
    }
    final_guest_ids = tuple(
        sorted(source_guest_ids - excluded_guest_ids - without_phone_guest_ids)
    )

    provider_counts = {"telegram": 0, "vk": 0, "max": 0}
    historical_guests = 0
    for guest_id in final_guest_ids:
        row = selected_row_by_guest_id[guest_id]
        for provider in set(row.providers):
            if provider in provider_counts:
                provider_counts[provider] += 1
        if set(row.channel_modes) & _HISTORICAL_CHANNEL_MODES:
            historical_guests += 1

    source_total = len(source_guest_ids)
    final_total = len(final_guest_ids)
    final_share = round((final_total * 100 / source_total), 2) if source_total else 0.0
    blockers = list(_mailing_state_blockers(mailing))
    warnings: list[str] = []

    if not matched_phones:
        blockers.append(
            "Ни один телефон из Excel не найден в базе. Аудитория не создана."
        )
    if source_total == 0:
        blockers.append(
            "Для выбранных ботов и группы каналов нет технически доставляемых гостей."
        )
    if final_total == 0 and source_total > 0:
        blockers.append("После применения исключений итоговая аудитория пуста.")
    if historical_guests and mailing.button_set != InteractionButtonSet.NONE:
        blockers.append(
            "Выбранная аудитория содержит исторические Telegram-маршруты, "
            "которые нельзя использовать в рассылке с кнопками."
        )

    if excel.invalid_rows:
        warnings.append(
            f"Строк с некорректным или пустым телефоном: {excel.invalid_rows}."
        )
    if excel.duplicate_rows:
        warnings.append(f"Повторных строк телефона: {excel.duplicate_rows}.")
    unmatched_phones = len(exclusion_phones - matched_phones)
    if unmatched_phones:
        warnings.append(f"Телефонов, не найденных в базе: {unmatched_phones}.")
    if without_phone_guest_ids:
        warnings.append(
            "Технически доставляемые гости без нормализуемого телефона не включены: "
            f"{len(without_phone_guest_ids)}."
        )
    warning_large_audience = bool(
        source_total and final_total * 100 >= source_total * REVERSE_IMPORT_WARNING_PERCENT
    )
    if warning_large_audience:
        warnings.append(
            f"Итоговая аудитория составляет {final_share:.2f}% исходной "
            "технически доставляемой аудитории. Проверьте режим и Excel."
        )

    campaign_fingerprint = _campaign_fingerprint(mailing, configured_bots)
    audience_digest = _audience_digest(
        selected_rows=selected_rows,
        normalized_phone_by_guest_id=normalized_phone_by_guest_id,
        exclusion_phones=exclusion_phones,
    )
    digest_payload = {
        "mailing_id": int(mailing.id),
        "audience_group": normalized_group,
        "file_sha256": excel.file_sha256,
        "campaign_fingerprint": campaign_fingerprint,
        "audience_digest": audience_digest,
        "matched_phones": len(matched_phones),
        "matched_guest_records": matched_guest_records,
        "source_deliverable_guests": source_total,
        "excluded_guest_records": len(excluded_guest_ids),
        "final_recipients": final_total,
        "guests_without_normalized_phone": len(without_phone_guest_ids),
    }
    calculation_digest = _stable_digest(digest_payload)

    return ReverseMailingImportCalculation(
        mailing_id=int(mailing.id),
        audience_group=normalized_group,
        audience_group_label=MAILING_IMPORT_AUDIENCE_LABELS[normalized_group],
        file_sha256=excel.file_sha256,
        total_rows=excel.total_rows,
        valid_rows=excel.valid_rows,
        invalid_rows=excel.invalid_rows,
        duplicate_rows=excel.duplicate_rows,
        unique_phones=len(exclusion_phones),
        matched_phones=len(matched_phones),
        unmatched_phones=unmatched_phones,
        matched_guest_records=matched_guest_records,
        source_deliverable_guests=source_total,
        excluded_guest_records=len(excluded_guest_ids),
        final_recipients=final_total,
        final_share_percent=final_share,
        guests_without_normalized_phone=len(without_phone_guest_ids),
        provider_telegram=provider_counts["telegram"],
        provider_vk=provider_counts["vk"],
        provider_max=provider_counts["max"],
        historical_guests=historical_guests,
        warning_large_audience=warning_large_audience,
        blockers=tuple(dict.fromkeys(blockers)),
        warnings=tuple(dict.fromkeys(warnings)),
        campaign_fingerprint=campaign_fingerprint,
        audience_digest=audience_digest,
        calculation_digest=calculation_digest,
        final_guest_ids=final_guest_ids,
    )


def create_reverse_import_preview_token(
    calculation: ReverseMailingImportCalculation,
) -> str:
    """Подписывает только отпечатки и агрегаты предварительного расчёта."""

    if not calculation.can_confirm:
        raise ReverseMailingImportError(
            "Предварительный расчёт содержит блокировки и не может быть подтверждён."
        )
    return signing.dumps(
        {
            "version": 1,
            "mailing_id": calculation.mailing_id,
            "operation": MAILING_IMPORT_OPERATION_EXCLUDE,
            "audience_group": calculation.audience_group,
            "file_sha256": calculation.file_sha256,
            "campaign_fingerprint": calculation.campaign_fingerprint,
            "audience_digest": calculation.audience_digest,
            "calculation_digest": calculation.calculation_digest,
        },
        salt=REVERSE_IMPORT_SIGNING_SALT,
        compress=True,
    )


def confirm_reverse_mailing_import(
    *,
    mailing_id: int,
    excel: ReverseExcelAnalysis,
    audience_group: str,
    preview_token: str,
) -> tuple[Mailing, ReverseMailingImportCalculation, int]:
    """Повторно рассчитывает и атомарно создаёт подтверждённую аудиторию."""

    token_payload = _load_preview_token(preview_token)
    normalized_group = normalize_mailing_import_audience(audience_group)
    expected_basics = {
        "version": 1,
        "mailing_id": int(mailing_id),
        "operation": MAILING_IMPORT_OPERATION_EXCLUDE,
        "audience_group": normalized_group,
        "file_sha256": excel.file_sha256,
    }
    for key, expected_value in expected_basics.items():
        if token_payload.get(key) != expected_value:
            raise ReverseMailingImportPreviewMismatch(
                "Файл или параметры импорта изменились. Выполните предварительный расчёт заново."
            )

    with transaction.atomic():
        try:
            mailing = Mailing.objects.select_for_update().get(pk=mailing_id)
        except Mailing.DoesNotExist as exc:
            raise ReverseMailingImportError("Кампания не найдена.") from exc

        state_blockers = _mailing_state_blockers(mailing)
        if state_blockers:
            raise ReverseMailingImportError(" ".join(state_blockers))

        calculation = build_reverse_mailing_import_calculation(
            mailing=mailing,
            excel=excel,
            audience_group=normalized_group,
        )
        if calculation.blockers:
            raise ReverseMailingImportError(" ".join(calculation.blockers))

        for key, actual_value in (
            ("campaign_fingerprint", calculation.campaign_fingerprint),
            ("audience_digest", calculation.audience_digest),
            ("calculation_digest", calculation.calculation_digest),
        ):
            if token_payload.get(key) != actual_value:
                raise ReverseMailingImportPreviewMismatch(
                    "Настройки или доступная аудитория изменились. "
                    "Выполните предварительный расчёт заново."
                )

        created_count = _create_mailing_guests(mailing, calculation.final_guest_ids)
        confirmed_at = timezone.now()
        source_snapshot = dict(mailing.source_filter_snapshot or {})
        source_snapshot["mailing_excel_import"] = {
            "operation": MAILING_IMPORT_OPERATION_EXCLUDE,
            "audience_group": calculation.audience_group,
            "file_sha256": calculation.file_sha256,
            "source_deliverable_guests": calculation.source_deliverable_guests,
            "excluded_guest_records": calculation.excluded_guest_records,
            "final_recipients": calculation.final_recipients,
            "final_share_percent": calculation.final_share_percent,
            "guests_without_normalized_phone": calculation.guests_without_normalized_phone,
            "matched_phones": calculation.matched_phones,
            "matched_guest_records": calculation.matched_guest_records,
            "confirmed_at": confirmed_at.isoformat(),
            "result": "confirmed",
        }
        mailing.source_filter_snapshot = source_snapshot
        mailing.save(update_fields=["source_filter_snapshot"])

    return mailing, calculation, created_count


def _create_mailing_guests(mailing: Mailing, guest_ids: Iterable[int]) -> int:
    """Создаёт строки аудитории пакетами, не удерживая весь набор объектов."""

    now = timezone.now()
    template_text = mailing.template.message_text
    batch: list[MailingGuest] = []
    created_count = 0
    guests = Guest.objects.filter(id__in=tuple(guest_ids)).order_by("id")
    for guest in guests.iterator(chunk_size=1000):
        batch.append(
            MailingGuest(
                mailing=mailing,
                guest=guest,
                phone=guest.phone,
                email=guest.email,
                text_mailing_list=render_message_for_guest(template_text, guest),
                scheduled_datetime=mailing.scheduled_time_begin,
                status=MailingGuest.Status.PLANNED,
                external_id=None,
                created_at=now,
            )
        )
        if len(batch) >= 500:
            MailingGuest.objects.bulk_create(batch, batch_size=500)
            created_count += len(batch)
            batch.clear()
    if batch:
        MailingGuest.objects.bulk_create(batch, batch_size=500)
        created_count += len(batch)
    return created_count


def _mailing_state_blockers(mailing: Mailing) -> tuple[str, ...]:
    blockers: list[str] = []
    if mailing.is_active:
        blockers.append("Обратный импорт разрешён только для выключенной кампании.")
    if mailing.is_archived:
        blockers.append("Обратный импорт недоступен для архивной кампании.")
    if MailingGuest.objects.filter(mailing=mailing).exists():
        blockers.append("Обратный импорт разрешён только для кампании без аудитории.")
    if DispatchTask.objects.filter(mailing_guest__mailing=mailing).exists():
        blockers.append("В кампании уже существуют задачи отправки.")
    return tuple(blockers)


def _collect_delivery_candidate_guest_ids(
    configured_bots: list[dict[str, Any]],
) -> tuple[int, ...]:
    """Собирает доказуемый наднабор гостей с потенциальным выбранным каналом."""

    active_bot_ids = {
        int(row["id"])
        for row in configured_bots
        if bool(row.get("is_active"))
    }
    if not active_bot_ids:
        return ()

    candidate_ids = {
        int(guest_id)
        for guest_id in (
            GuestBotBinding.objects.filter(
                bot_id__in=active_bot_ids,
                is_active=True,
                bot__is_active=True,
            )
            .exclude(external_chat_id__isnull=True)
            .exclude(external_chat_id="")
            .values_list("guest_id", flat=True)
            .distinct()
        )
    }
    active_telegram_bot_ids = {
        int(row["id"])
        for row in configured_bots
        if bool(row.get("is_active"))
        and str(row.get("provider_type") or "") == BotProfile.ProviderType.TELEGRAM
    }
    if active_telegram_bot_ids:
        candidate_ids.update(
            int(guest_id)
            for guest_id in (
                HistoricalTelegramChannel.objects.filter(
                    bot_profile_id__in=active_telegram_bot_ids,
                    bot_profile__is_active=True,
                    delivery_state=HistoricalTelegramChannel.DeliveryState.SENDABLE,
                )
                .exclude(telegram_chat_id="")
                .values_list("guest_id", flat=True)
                .distinct()
            )
        )
        candidate_ids.update(
            int(guest_id)
            for guest_id in (
                VtelemaxRecipientChannel.objects.filter(
                    platform=VtelemaxRecipientChannel.Platform.TELEGRAM,
                    is_registered=True,
                    notifications_allowed=True,
                )
                .exclude(guest_id__isnull=True)
                .exclude(external_id__isnull=True)
                .exclude(external_id="")
                .values_list("guest_id", flat=True)
                .distinct()
            )
        )
    return tuple(sorted(candidate_ids))


def _row_belongs_to_group(row: MailingDeliveryRow, audience_group: str) -> bool:
    modes = set(row.channel_modes)
    if audience_group == MAILING_IMPORT_AUDIENCE_HISTORICAL_TELEGRAM:
        return bool(modes & _HISTORICAL_CHANNEL_MODES)
    if audience_group == MAILING_IMPORT_AUDIENCE_ALL_SENDABLE:
        return True
    return CHANNEL_MODE_BINDING in modes


def _campaign_fingerprint(mailing: Mailing, configured_bots: list[dict[str, Any]]) -> str:
    template = mailing.template
    payload = {
        "mailing_id": int(mailing.id),
        "updated_at": _serialize_value(mailing.updated_at),
        "is_active": bool(mailing.is_active),
        "is_archived": bool(mailing.is_archived),
        "template_id": int(mailing.template_id),
        "template_updated_at": _serialize_value(getattr(template, "updated_at", None)),
        "template_text_sha256": hashlib.sha256(
            str(template.message_text or "").encode("utf-8")
        ).hexdigest(),
        "scheduled_time_begin": _serialize_value(mailing.scheduled_time_begin),
        "scheduled_time_end": _serialize_value(mailing.scheduled_time_end),
        "send_window_begin": _serialize_value(mailing.send_window_begin),
        "send_window_end": _serialize_value(mailing.send_window_end),
        "target_mode": mailing.target_mode,
        "queue_priority": mailing.queue_priority,
        "button_set": mailing.button_set,
        "tracked_link_destination_id": mailing.tracked_link_destination_id,
        "coupon_series": mailing.coupon_series,
        "source_filter_snapshot": mailing.source_filter_snapshot or {},
        "configured_bots": configured_bots,
    }
    return _stable_digest(payload)


def _audience_digest(
    *,
    selected_rows: Iterable[MailingDeliveryRow],
    normalized_phone_by_guest_id: dict[int, str],
    exclusion_phones: set[str],
) -> str:
    payload = []
    for row in selected_rows:
        phone = normalized_phone_by_guest_id.get(int(row.guest_id))
        payload.append(
            {
                "guest_id": int(row.guest_id),
                "target_count": int(row.target_count),
                "providers": list(row.providers),
                "bot_profile_ids": list(row.bot_profile_ids),
                "channel_modes": list(row.channel_modes),
                "phone_normalizable": phone is not None,
                "excluded_by_excel": phone in exclusion_phones,
            }
        )
    return _stable_digest(payload)


def _load_preview_token(token: str) -> dict[str, Any]:
    if not str(token or "").strip():
        raise ReverseMailingImportPreviewMismatch(
            "Отсутствует подтверждённый предварительный расчёт. Выполните его заново."
        )
    try:
        payload = signing.loads(
            token,
            salt=REVERSE_IMPORT_SIGNING_SALT,
            max_age=REVERSE_IMPORT_TOKEN_MAX_AGE_SECONDS,
        )
    except signing.SignatureExpired as exc:
        raise ReverseMailingImportPreviewMismatch(
            "Срок действия предварительного расчёта истёк. Выполните его заново."
        ) from exc
    except signing.BadSignature as exc:
        raise ReverseMailingImportPreviewMismatch(
            "Предварительный расчёт повреждён. Выполните его заново."
        ) from exc
    if not isinstance(payload, dict):
        raise ReverseMailingImportPreviewMismatch(
            "Предварительный расчёт имеет неверный формат. Выполните его заново."
        )
    return payload


def _hash_uploaded_file(file_obj) -> str:
    digest = hashlib.sha256()
    try:
        file_obj.seek(0)
        while True:
            chunk = file_obj.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    except (AttributeError, OSError) as exc:
        raise ReverseMailingImportError("Не удалось прочитать загруженный Excel.") from exc
    finally:
        try:
            file_obj.seek(0)
        except (AttributeError, OSError):
            pass
    return digest.hexdigest()


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _serialize_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _stable_digest(payload: Any) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
